from openpyxl import load_workbook

def add_data_to_sheet(cell_location,test_cases,successful_checks):
    file_path="Automatic.xlsx"
    wb = load_workbook(filename = file_path)
    ws=wb['TESZTESETEK']
    # send the num of successful check and all checks to the file
    ws[cell_location] = f"{test_cases}/{successful_checks}"
    wb.save(file_path)
